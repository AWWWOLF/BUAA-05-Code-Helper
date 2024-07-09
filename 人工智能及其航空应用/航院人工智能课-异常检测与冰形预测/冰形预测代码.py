# 异常数据检测.py
# ---------------------------------------------------------------------------
import numpy as np
from scipy.stats import norm
import re
if __name__ == "__main__":
# 假设训练集和测试集都是m x n的二维数组
    X1 = []
    X2 = []
    yval = []
    filename = ["naca_0.5_257_250.txt", "naca_0.5_261_250.txt", "naca_0.5_263_250.txt", "naca_0.5_265_250.txt"]
    for i in range(0, 3):
        with open(filename[i], 'r') as file_to_read:  # 读取训练集的数据
            while True:
                lines = file_to_read.readline()
                if not lines:
                    break
                x1_tmp, x2_tmp = [float(i) for i in lines.split(" ")]
                X1.append(x1_tmp)
                X2.append(x2_tmp)
    X1 = np.array(X1)
    X1 = X1[np.newaxis, :]
    X2 = np.array(X2)
    X2 = X2[np.newaxis, :]
    train_set = np.concatenate((X1.T, X2.T), axis=1)
    Z1 = []
    Z2 = []
    with open('naca_0.5_259_250.txt', 'r') as file_to_read:  # 读取训练集的数据
        while True:
            lines = file_to_read.readline()
            lines = re.sub("[a-d]", "7", lines)  # 将匹配到的一些字母替换为7
            if not lines:
                break
            z1_tmp, z2_tmp = [float(i) for i in lines.split(" ")]
            Z1.append(z1_tmp)
            Z2.append(z2_tmp)
    Z1 = np.array(Z1)
    Z1 = Z1[np.newaxis, :]
    Z2 = np.array(Z2)
    Z2 = Z2[np.newaxis, :]
    test_set = np.concatenate((Z1.T, Z2.T), axis=1)
    # 定义一个阈值，低于这个值的样本被认为是异常的
    threshold = 0.07
    # 对于每一个特征，计算训练集中该特征的均值和方差
    means = np.mean(train_set, axis=0)
    stds = np.std(train_set, axis=0)
    # 对于测试集中的每一个样本，计算它在每一个特征上的高斯PDF的值
    pdfs = norm.pdf(test_set, means, stds)
    # 找出低于阈值的样本，让它等于平均值
    for i in range(0, len(pdfs)):
        if (pdfs[i, 0] <= threshold) or (pdfs[i, 1] <= threshold):
            test_set[i, :] = (test_set[i-1, :]+test_set[i+1, :])/2
    # 输出新的测试集
    print(test_set)
    np.savetxt("correct.txt_naca_0.5_259_250.txt", test_set, fmt="%f", delimiter=" ")
    
# ----------------------------------------------------------------------------
    
# txt转excel-1.py
#-----------------------------------------------------------------------------
import openpyxl
import numpy as np
if __name__ == "__main__":
    X1 = []
    X2 = []
    X = []
    Y = []
    filename = ["naca_0.5_257_250.txt", "naca_0.5_261_250.txt", "naca_0.5_263_250.txt", "naca_0.5_265_250.txt", "correct.txt_naca_0.5_259_250.txt"]
    with open(filename[0], 'r') as file_to_read:  # 读取训练集的数据
        while True:
            lines = file_to_read.readline()
            if not lines:
                break
            x1_tmp, x2_tmp = [float(i) for i in lines.split(" ")]
            Y.append(x2_tmp)
            X.append(x1_tmp)
        X = np.array(X)
        Y = np.array(Y)
    for i in range(1, 5):
        X2 = []
        X1 = []
        with open(filename[i], 'r') as file_to_read:  # 读取训练集的数据
            while True:
                lines = file_to_read.readline()
                if not lines:
                    break
                x1_tmp, x2_tmp = [float(i) for i in lines.split(" ")]
                X2.append(x2_tmp)
                X1.append(x1_tmp)
        X1 = np.array(X1)
        X = np.vstack((X, X1))
        X2 = np.array(X2)
        Y = np.vstack((Y, X2))

    # 创建一个workbook对象
    wb1 = openpyxl.Workbook()
    # 获取活动的worksheet对象
    ws = wb1.active
    # 获取二维数组的行数和列数
    Xrows = len(X)
    Xcols = len(X[0])

    # 遍历数组，将元素写入到worksheet中
    for i in range(X.shape[0]):
        for j in range(X.shape[1]):
            ws.cell(row=j + 1, column=i + 1, value=X[i, j])
    # 保存workbook为xlwt文件'
    wb1.save("exampleX.xlsx")
    # 关闭workbook
    wb1.close()
    
    wb2 = openpyxl.Workbook()
    # 获取活动的worksheet对象
    ws = wb2.active
    # 获取二维数组的行数和列数
    Yrows = len(Y)
    Ycols = len(Y[0])

    # 遍历数组，将元素写入到worksheet中
    for i in range(Y.shape[0]):
        for j in range(Y.shape[1]):
            ws.cell(row=j + 1, column=i + 1, value=Y[i, j])
    # 保存workbook为xlwt文件'
    wb2.save("exampleY.xlsx")
    # 关闭workbook
    wb2.close()
# ----------------------------------------------------------------------------

# POD.py
# ----------------------------------------------------------------------------
import xlrd
import numpy as np
from sklearn.decomposition import PCA
import pandas as pd
import joblib


# 打开数据文件，读入xls格式数据，InputPath为样本数据路径，读入的数据维度为x*y,其中x为一个样本的节点数，y为总样本个数
def ReadSampleData(InputPath, x, y):
    workbook = xlrd.open_workbook(InputPath)
    sheet1 = workbook.sheets()[0]
    # 每一行为一个维度，将每一个维度的数据提取出来
    data = np.zeros(shape=(0, y))
    for i in range(0, x):
        rows = sheet1.row_values(i)
        data = np.append(data, [rows], axis=0)
    print(data)
    return data


# 进行数据降维，data为导入的样本数据, n为降维后的数据维度
def POD(data, n):

    data = data.astype(float)  # 转换data内数据由string类型至float类型
    cov_f = np.cov(data)
    print("协方差矩阵：", cov_f)

    eigenvalue, featurevector = np.linalg.eig(cov_f)
    print("特征值：", eigenvalue)
    print("特征向量：", featurevector)

    data = data.T
    pca = PCA(n_components=n)
    pca.fit(data)
    data_afterPOD = pca.transform(data)
    print("降维后的数据：", data_afterPOD)

    # 降维后的各主成分的方差值占总方差值的比例，即方差贡献率
    print("方差贡献率：", pca.explained_variance_ratio_)

    # 总方差贡献率
    print("总方差贡献率", sum(pca.explained_variance_ratio_))

    # 降维后的各主成分的方差值
    print("主成分方差：", pca.explained_variance_)
    return pca, data_afterPOD  # pca为降维模型，data_afterPOD为降维后的数据


# 降维后的数据写入EXCEL, data_afterPOD为降维后的数据，OutputPath为降维后数据保存路径
def Output(data_afterPOD, OutputPath):
    pdData = pd.DataFrame(data_afterPOD)
    writer = pd.ExcelWriter(OutputPath)  # 写入Excel文件
    pdData.to_excel(writer, '1')  # ‘1’是写入excel的sheet名
    writer.save()
    writer.close()
    print("Output Done")


# 保存POD模型
def SavePodModel(Pod_model, OutputPath):
    joblib.dump(Pod_model, OutputPath)
    print("Model Saved")


# POD程序运行,需要自行修改以下代码
SampleDataInputPath = r'exampleX.xlsx'
data = ReadSampleData(SampleDataInputPath, 320, 5)
pca, data_afterPOD = POD(data, 1)
OutputPath1 = r'exampleX_afterPOD.xlsx'
Output(data_afterPOD, OutputPath1)
OutputPath2 = r'POD_X.model'
SavePodModel(pca, OutputPath2)

SampleDataInputPath = r'exampleY.xlsx'
data = ReadSampleData(SampleDataInputPath, 320, 5)
pca, data_afterPOD = POD(data, 1)
OutputPath1 = r'exampleY_afterPOD.xlsx'
Output(data_afterPOD, OutputPath1)
OutputPath2 = r'POD_Y.model'
SavePodModel(pca, OutputPath2)
# ----------------------------------------------------------------------------

# SVM.py
# ----------------------------------------------------------------------------
from sklearn import svm
from sklearn.preprocessing import StandardScaler, RobustScaler
from sklearn.svm import SVR
from sklearn.model_selection import train_test_split
import matplotlib.pyplot as plt
from sklearn.metrics import mean_squared_error, r2_score
import pandas as pd
import openpyxl
import numpy as np

if __name__ == "__main__":
    # 读取xlsx文件，假设文件名为data.xlsx，工作表名为Sheet1，跳过第一行和第一列
    data1 = pd.read_excel("exampleY_afterPOD.xlsx", sheet_name="1", skiprows=0, usecols="B")
    # 转换为数组
    data1 = data1.values
    data4 = pd.read_excel("exampleX_afterPOD.xlsx", sheet_name="1", skiprows=0, usecols="B")
    data4 = data4.values
    X = np.array([257,259,261,263,265])
    Xnew = X.reshape(-1, 1)
    print(Xnew)
    rbX = RobustScaler()
    rbY1 = RobustScaler()
    rbY4 = RobustScaler()
    X_train = rbX.fit_transform(Xnew)
    data1_temp = rbY1.fit_transform(data1.reshape(-1, 1))
    data4_temp = rbY4.fit_transform(data4.reshape(-1, 1))
    data1 = np.ravel(data1_temp)
    data4 = np.ravel(data4_temp)
    print(data1)

    # 创建一个SVR回归对象，使用高斯核，正则化系数为0.1
    model1 = svm.SVR(kernel='rbf', C=100, gamma=0.01)
    # 拟合SVR回归对象
    model1.fit(Xnew, np.array(data1))
    # 预测SVR回归对象
    y_pred1 = model1.predict(Xnew).reshape(-1, 1)
    model4 = svm.SVR(kernel='rbf', C=100, gamma=0.01)
    model4.fit(Xnew, data4)
    # 预测SVR回归对象
    x_pred4 = model4.predict(Xnew).reshape(-1, 1)
    X_train = rbX.inverse_transform(X_train)
    data1 = rbY1.inverse_transform(data1_temp)
    data4 = rbY4.inverse_transform(data4_temp)
    y_pred1 = rbY1.inverse_transform(y_pred1).ravel()
    x_pred4 = rbY4.inverse_transform(x_pred4).ravel()
    ArrayY = y_pred1.T
    ArrayX = x_pred4.T


    plt.plot(X_train, data1, 'red', label="YT")
    plt.plot(X_train, y_pred1, 'blue', label="YP")
    plt.legend()
    plt.plot(X_train, data4, 'black', label="XT")
    plt.plot(X_train, x_pred4, 'green', label="XP")
    plt.legend()
    plt.show()

    # 评估SVR回归对象
    mse1 = mean_squared_error(data1, y_pred1)
    mse4 = mean_squared_error(data4, x_pred4)

    # 创建一个workbook对象
    wb1 = openpyxl.Workbook()
    # 获取活动的worksheet对象
    ws1 = wb1.active
    # 获取二维数组的行数和列数
    rows1 = len(ArrayX)
    # 遍历数组，将元素写入到worksheet中
    for i in range(len(ArrayX)):
        ws1.cell(row=i + 1, column=1, value=ArrayX[i])
    # 保存workbook为xlwt文件
    wb1.save("svmpredict_X.xlsx")
    # 关闭workbook
    wb1.close()

    # 创建一个workbook对象
    wb1 = openpyxl.Workbook()
    # 获取活动的worksheet对象
    ws1 = wb1.active
    # 获取二维数组的行数和列数
    rows1 = len(ArrayY)
    # 遍历数组，将元素写入到worksheet中
    for i in range(len(ArrayY)):
        ws1.cell(row=i + 1, column=1, value=ArrayY[i])
    # 保存workbook为xlwt文件
    wb1.save("svmpredict_Y.xlsx")
    # 关闭workbook
    wb1.close()


# ----------------------------------------------------------------------------

# InversePod.py
# ----------------------------------------------------------------------------
import pandas as pd
import joblib
import openpyxl
import numpy as np

# 导入POD模型
def LoadPodModel(InputPath):
    Pod_model = joblib.load(InputPath)
    print("Model Loaded")
    return Pod_model


# 进行逆POD，返回对应的完整数据
def InversePod(Pod_model, NN_predict):
    result = Pod_model.inverse_transform(NN_predict)
    print("预测结果（逆pod下）：", result)
    return result


# 数据写入EXCEL, data为数据，Output为降维后数据保存路径
def Output(data, OutputPath):
    pdData = pd.DataFrame(data)
    writer = pd.ExcelWriter(OutputPath)  # 写入Excel文件
    pdData.to_excel(writer, '1')  # ‘1’是写入excel的sheet名
    writer.save()
    writer.close()
    print("Output Done")


# 以下需自行修改
# 打开Excel文件，返回一个Workbook对象
wb = openpyxl.load_workbook("svmpredict_X.xlsx")
# 获取指定的工作表，返回一个Worksheet对象
ws = wb["Sheet"]
# 获取Worksheet对象中的所有单元格的值，返回一个生成器
values = ws.values
# 转换为一个numpy数组
array = np.array(list(values))
# 神经网络预测值
SVM_predict = array
# POD模型存储位置
Pod_model = LoadPodModel(r'POD_X.model')
# 逆POD输出结果
result = InversePod(Pod_model, SVM_predict)
# 输出位置
OutputPath = r'SVM_predict_resultX.xlsx'
Output(result, OutputPath)
# 以下需自行修改
# 打开Excel文件，返回一个Workbook对象
wb = openpyxl.load_workbook("svmpredict_Y.xlsx")
# 获取指定的工作表，返回一个Worksheet对象
ws = wb["Sheet"]
# 获取Worksheet对象中的所有单元格的值，返回一个生成器
values = ws.values
# 转换为一个numpy数组
array = np.array(list(values))
# 神经网络预测值
SVM_predict = array
# POD模型存储位置
Pod_model = LoadPodModel(r'POD_Y.model')
# 逆POD输出结果
result = InversePod(Pod_model, SVM_predict)
# 输出位置
OutputPath = r'SVM_predict_resultY.xlsx'
Output(result, OutputPath)

# ----------------------------------------------------------------------------
# 绘图
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np

x = pd.read_excel("SVM_predict_resultX.xlsx", sheet_name="1", skiprows=0, usecols=lambda column: column not in ['A'])
x = np.array(x)
y = pd.read_excel("SVM_predict_resultY.xlsx", sheet_name="1", skiprows=0, usecols=lambda column: column not in ['A'])
y = np.array(y)
# 使用循环绘制5条折线，每条折线用不同的颜色和标记表示
colors = ["red", "green", "blue", "orange", "purple"]
for i in range(5):
    for j in range(1, x.shape[1]):  # 按照不同聚类用不同的颜色打印点
        plt.scatter(x[i, j], y[i, j], marker='o', s=1, c=colors[i])
# 显示图形
plt.show()